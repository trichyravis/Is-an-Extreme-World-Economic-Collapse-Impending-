
"""
Macro Risk Dashboard — Streamlit app
====================================
Interactive front-end for the model from
"Is an Extreme World Economic Collapse Impending?"

How to run
----------
    pip install -r requirements.txt
    streamlit run streamlit_app.py

The app exposes the same four modules as risk_model.py / build_dashboard.py
but with live sidebar controls, so you can tweak indicators, debt parameters,
scenario shocks, and Monte Carlo settings and see results update immediately.

Tabs
----
  Overview     — headline scorecard and snapshot
  SRI          — composite systemic risk index, indicator z-scores
  Debt         — debt sustainability dynamics under r-g-pb scenarios
  Monte Carlo  — copula-based factor stress test, VaR / CVaR
  Scenarios    — single-shot Base / Slow Burn / Energy Shock + Trade War
  Download     — generate and download an Excel dashboard
"""

from __future__ import annotations

import io
import json
from dataclasses import dataclass
from pathlib import Path

import numpy as np
import pandas as pd
import matplotlib.pyplot as plt
from numpy.linalg import eigh

import streamlit as st

from openpyxl import Workbook
from openpyxl.styles import Alignment, Border, Font, PatternFill, Side
from openpyxl.utils import get_column_letter
from openpyxl.formatting.rule import ColorScaleRule


# ===========================================================================
# Page config — Mountain Path Academy aesthetic
# ===========================================================================
st.set_page_config(
    page_title="Macro Risk Dashboard · The Mountain Path Academy",
    page_icon="⛰️",
    layout="wide",
    initial_sidebar_state="expanded",
)

# ---------- Brand palette ----------
NAVY        = "#1B3A5B"   # primary — mountain twilight
NAVY_DEEP   = "#0F2540"   # peak shadow, for headings
GOLD        = "#C9A961"   # sunrise gold, accent
GOLD_DARK   = "#A8893F"   # hover / line accent
PARCHMENT   = "#FAF7F1"   # page background
PARCHMENT_2 = "#F2EAD3"   # secondary background
CHARCOAL    = "#2C3E50"   # body text
SLATE       = "#5D6D7E"   # muted text
MOUNTAIN_GREEN = "#5B7C5A"  # subtle eco accent
ALERT_RED   = "#A33B2A"   # crisis / drawdown
ALERT_AMBER = "#D08B2C"   # stress

# Backward-compatible name used by older chart code
ACCENT = NAVY

# ---------- Custom CSS ----------
_CSS = f"""
<style>
@import url('https://fonts.googleapis.com/css2?family=Playfair+Display:wght@500;600;700&family=Inter:wght@300;400;500;600;700&display=swap');

html, body, [class*="css"], .stApp {{
    font-family: 'Inter', -apple-system, BlinkMacSystemFont, sans-serif !important;
    color: {CHARCOAL};
}}
.stApp {{
    background: {PARCHMENT};
}}
h1, h2, h3, h4 {{
    font-family: 'Playfair Display', Georgia, serif !important;
    color: {NAVY_DEEP} !important;
    letter-spacing: -0.01em;
}}
h1 {{ font-weight: 700; }}
h2, h3 {{ font-weight: 600; }}

/* Sidebar */
section[data-testid="stSidebar"] {{
    background: linear-gradient(180deg, {NAVY_DEEP} 0%, {NAVY} 100%);
}}
section[data-testid="stSidebar"] * {{ color: #EFE7D2 !important; }}
section[data-testid="stSidebar"] h1,
section[data-testid="stSidebar"] h2,
section[data-testid="stSidebar"] h3 {{
    color: {GOLD} !important;
    font-family: 'Playfair Display', Georgia, serif !important;
}}
section[data-testid="stSidebar"] [data-baseweb="slider"] [role="slider"] {{
    background-color: {GOLD} !important;
}}
section[data-testid="stSidebar"] .stExpander {{
    background: rgba(255,255,255,0.04);
    border: 1px solid rgba(201,169,97,0.18);
    border-radius: 8px;
}}

/* --- Sidebar inputs: dark navy text on white field, gold accents --- */
section[data-testid="stSidebar"] input,
section[data-testid="stSidebar"] textarea,
section[data-testid="stSidebar"] [data-baseweb="input"] input,
section[data-testid="stSidebar"] [data-baseweb="select"] input,
section[data-testid="stSidebar"] [data-baseweb="select"] div,
section[data-testid="stSidebar"] [data-baseweb="select"] span,
section[data-testid="stSidebar"] [data-baseweb="input"] *,
section[data-testid="stSidebar"] [data-testid="stNumberInput"] input,
section[data-testid="stSidebar"] [data-testid="stTextInput"] input,
section[data-testid="stSidebar"] [data-testid="stSelectbox"] * {{
    color: {NAVY_DEEP} !important;
    font-weight: 500;
    -webkit-text-fill-color: {NAVY_DEEP} !important;  /* Safari/Chrome override */
    background-color: #FFFFFF !important;
}}
/* Slider current-value chip + caption — keep these readable on dark bg */
section[data-testid="stSidebar"] [data-baseweb="slider"] [data-testid="stTickBarMin"],
section[data-testid="stSidebar"] [data-baseweb="slider"] [data-testid="stTickBarMax"] {{
    color: {GOLD} !important;
    -webkit-text-fill-color: {GOLD} !important;
    background: transparent !important;
}}
/* Number-input +/- step buttons */
section[data-testid="stSidebar"] [data-testid="stNumberInput"] button {{
    color: {NAVY_DEEP} !important;
    -webkit-text-fill-color: {NAVY_DEEP} !important;
    background-color: {PARCHMENT_2} !important;
    border-color: {GOLD} !important;
}}
section[data-testid="stSidebar"] [data-testid="stNumberInput"] button:hover {{
    background-color: {GOLD} !important;
    color: {NAVY_DEEP} !important;
}}
/* Caret + placeholder colour */
section[data-testid="stSidebar"] input::placeholder {{
    color: {SLATE} !important;
    -webkit-text-fill-color: {SLATE} !important;
}}
section[data-testid="stSidebar"] input {{ caret-color: {NAVY_DEEP}; }}

/* --- Sidebar expander headers (Global settings, Current indicator readings, etc.)
       Default + hover + focus states all preserved in high-contrast gold/cream. --- */
section[data-testid="stSidebar"] details > summary,
section[data-testid="stSidebar"] [data-testid="stExpander"] summary,
section[data-testid="stSidebar"] .streamlit-expanderHeader,
section[data-testid="stSidebar"] [data-testid="stExpander"] > details > summary,
section[data-testid="stSidebar"] [data-testid="stExpander"] [role="button"] {{
    color: {GOLD} !important;
    -webkit-text-fill-color: {GOLD} !important;
    background-color: rgba(255,255,255,0.04) !important;
    font-weight: 600;
    font-family: 'Playfair Display', Georgia, serif !important;
    letter-spacing: 0.01em;
    border-radius: 6px;
}}
section[data-testid="stSidebar"] details > summary *,
section[data-testid="stSidebar"] [data-testid="stExpander"] summary *,
section[data-testid="stSidebar"] .streamlit-expanderHeader * {{
    color: {GOLD} !important;
    -webkit-text-fill-color: {GOLD} !important;
}}
section[data-testid="stSidebar"] details > summary:hover,
section[data-testid="stSidebar"] [data-testid="stExpander"] summary:hover,
section[data-testid="stSidebar"] .streamlit-expanderHeader:hover,
section[data-testid="stSidebar"] [data-testid="stExpander"] [role="button"]:hover {{
    color: {PARCHMENT} !important;
    -webkit-text-fill-color: {PARCHMENT} !important;
    background-color: rgba(201,169,97,0.18) !important;
}}
section[data-testid="stSidebar"] details > summary:hover *,
section[data-testid="stSidebar"] [data-testid="stExpander"] summary:hover *,
section[data-testid="stSidebar"] .streamlit-expanderHeader:hover * {{
    color: {PARCHMENT} !important;
    -webkit-text-fill-color: {PARCHMENT} !important;
}}
/* Expander chevron icon */
section[data-testid="stSidebar"] [data-testid="stExpanderToggleIcon"] svg,
section[data-testid="stSidebar"] details summary svg {{
    fill: {GOLD} !important;
    color: {GOLD} !important;
}}
section[data-testid="stSidebar"] details > summary:hover svg,
section[data-testid="stSidebar"] .streamlit-expanderHeader:hover svg {{
    fill: {PARCHMENT} !important;
    color: {PARCHMENT} !important;
}}

/* --- Slider labels + the current-value bubble above the thumb --- */
section[data-testid="stSidebar"] [data-testid="stWidgetLabel"],
section[data-testid="stSidebar"] [data-testid="stWidgetLabel"] *,
section[data-testid="stSidebar"] label,
section[data-testid="stSidebar"] label * {{
    color: #EFE7D2 !important;
    -webkit-text-fill-color: #EFE7D2 !important;
    font-weight: 500;
}}
section[data-testid="stSidebar"] [data-baseweb="slider"] [data-testid="stThumbValue"],
section[data-testid="stSidebar"] [data-baseweb="slider"] [data-testid="stThumbValue"] *,
section[data-testid="stSidebar"] [data-baseweb="tooltip"],
section[data-testid="stSidebar"] [data-baseweb="tooltip"] * {{
    color: {NAVY_DEEP} !important;
    -webkit-text-fill-color: {NAVY_DEEP} !important;
    background-color: {GOLD} !important;
    font-weight: 600;
    border-radius: 4px;
}}
/* Slider min/max tick labels */
section[data-testid="stSidebar"] [data-baseweb="slider"] [data-testid="stTickBar"] *,
section[data-testid="stSidebar"] [data-baseweb="slider"] [data-testid="stTickBarMin"],
section[data-testid="stSidebar"] [data-baseweb="slider"] [data-testid="stTickBarMax"] {{
    color: rgba(239,231,210,0.7) !important;
    -webkit-text-fill-color: rgba(239,231,210,0.7) !important;
    font-size: 0.72rem;
}}

/* --- Markdown / captions / strong text inside the sidebar --- */
section[data-testid="stSidebar"] p,
section[data-testid="stSidebar"] li,
section[data-testid="stSidebar"] .stMarkdown,
section[data-testid="stSidebar"] .stMarkdown * {{
    color: #EFE7D2 !important;
    -webkit-text-fill-color: #EFE7D2 !important;
}}
section[data-testid="stSidebar"] strong,
section[data-testid="stSidebar"] b {{
    color: {GOLD} !important;
    -webkit-text-fill-color: {GOLD} !important;
}}

/* Tabs */
.stTabs [data-baseweb="tab-list"] {{
    gap: 0.25rem;
    border-bottom: 2px solid {GOLD};
}}
.stTabs [data-baseweb="tab"] {{
    background: transparent;
    color: {SLATE};
    font-weight: 500;
    padding: 0.5rem 1rem;
}}
.stTabs [aria-selected="true"] {{
    color: {NAVY_DEEP} !important;
    border-bottom: 3px solid {GOLD} !important;
    font-weight: 600;
}}

/* Metrics */
[data-testid="stMetric"] {{
    background: #FFFFFF;
    border: 1px solid {PARCHMENT_2};
    border-left: 4px solid {GOLD};
    border-radius: 6px;
    padding: 1rem 1.2rem;
    box-shadow: 0 1px 2px rgba(15,37,64,0.04);
}}
[data-testid="stMetricLabel"] {{
    color: {SLATE} !important;
    font-size: 0.78rem !important;
    text-transform: uppercase;
    letter-spacing: 0.06em;
    font-weight: 500;
}}
[data-testid="stMetricValue"] {{
    color: {NAVY_DEEP} !important;
    font-family: 'Playfair Display', Georgia, serif !important;
    font-weight: 700;
}}

/* Buttons */
.stDownloadButton button, .stButton button {{
    background: {NAVY} !important;
    color: {PARCHMENT} !important;
    border: 1px solid {GOLD} !important;
    font-weight: 600;
    border-radius: 4px;
    transition: all 0.15s ease;
}}
.stDownloadButton button:hover, .stButton button:hover {{
    background: {GOLD} !important;
    color: {NAVY_DEEP} !important;
    border-color: {GOLD_DARK} !important;
}}

/* Brand header strip */
.mpa-header {{
    background: linear-gradient(135deg, {NAVY_DEEP} 0%, {NAVY} 60%, {NAVY_DEEP} 100%);
    border-bottom: 3px solid {GOLD};
    padding: 1.4rem 1.6rem 1.2rem;
    margin: -1rem -1rem 1.4rem -1rem;
    color: {PARCHMENT};
    border-radius: 0 0 6px 6px;
}}
.mpa-header .eyebrow {{
    color: {GOLD};
    font-family: 'Inter', sans-serif;
    font-size: 0.72rem;
    text-transform: uppercase;
    letter-spacing: 0.18em;
    font-weight: 600;
}}
.mpa-header h1 {{
    color: {PARCHMENT} !important;
    margin: 0.2rem 0 0.2rem;
    font-size: 2.1rem;
    line-height: 1.15;
}}
.mpa-header .sub {{
    color: {GOLD};
    font-style: italic;
    font-family: 'Playfair Display', Georgia, serif;
    font-size: 1.05rem;
}}
.mpa-header .by {{
    color: rgba(255,255,255,0.65);
    font-size: 0.82rem;
    margin-top: 0.35rem;
}}

/* Sidebar profile card */
.mpa-profile {{
    text-align: center;
    padding: 0.8rem 0.4rem 0.4rem;
    border-bottom: 1px solid rgba(201,169,97,0.3);
    margin-bottom: 0.6rem;
}}
.mpa-profile .name {{
    color: {GOLD} !important;
    font-family: 'Playfair Display', Georgia, serif;
    font-size: 1.15rem;
    font-weight: 600;
}}
.mpa-profile .tagline {{
    color: rgba(239,231,210,0.75) !important;
    font-size: 0.78rem;
    font-style: italic;
    margin-bottom: 0.6rem;
}}
.mpa-profile a {{
    display: inline-block;
    margin: 0 0.35rem;
    padding: 0.35rem 0.6rem;
    border: 1px solid {GOLD};
    border-radius: 4px;
    text-decoration: none !important;
    font-size: 0.78rem;
    color: {GOLD} !important;
    transition: all 0.15s ease;
}}
.mpa-profile a:hover {{
    background: {GOLD};
    color: {NAVY_DEEP} !important;
}}

/* Footer */
.mpa-footer {{
    margin-top: 2rem;
    padding: 1.4rem 1rem 1rem;
    border-top: 2px solid {GOLD};
    background: {PARCHMENT_2};
    border-radius: 6px;
    text-align: center;
    color: {CHARCOAL};
}}
.mpa-footer .col {{
    display: inline-block;
    vertical-align: top;
    margin: 0 1.4rem;
    text-align: left;
}}
.mpa-footer h4 {{
    color: {NAVY_DEEP} !important;
    margin: 0 0 0.4rem;
    font-size: 0.95rem;
    font-family: 'Playfair Display', Georgia, serif !important;
}}
.mpa-footer a {{
    color: {NAVY} !important;
    text-decoration: none;
    font-weight: 500;
}}
.mpa-footer a:hover {{ color: {GOLD_DARK} !important; }}
.mpa-footer .disclaimer {{
    color: {SLATE};
    font-size: 0.78rem;
    font-style: italic;
    margin-top: 1rem;
}}

/* Dataframes — softer borders */
[data-testid="stDataFrame"] {{
    border: 1px solid {PARCHMENT_2};
    border-radius: 4px;
}}

/* Expander header */
.streamlit-expanderHeader {{
    font-weight: 600;
    color: {NAVY_DEEP};
}}

/* Hide Streamlit chrome */
#MainMenu {{ visibility: hidden; }}
footer {{ visibility: hidden; }}
header[data-testid="stHeader"] {{ background: transparent; }}
</style>
"""
st.markdown(_CSS, unsafe_allow_html=True)

# ---------- Matplotlib theme to match ----------
plt.rcParams.update({
    "figure.figsize": (10, 6),
    "figure.facecolor": PARCHMENT,
    "axes.facecolor":   "#FFFFFF",
    "axes.edgecolor":   SLATE,
    "axes.labelcolor":  CHARCOAL,
    "axes.titlecolor":  NAVY_DEEP,
    "axes.titleweight": "bold",
    "axes.titlesize":   12,
    "axes.spines.top":   False,
    "axes.spines.right": False,
    "axes.grid":         True,
    "grid.color":        PARCHMENT_2,
    "grid.alpha":        0.7,
    "xtick.color":       SLATE,
    "ytick.color":       SLATE,
    "font.family":       "DejaVu Sans",  # Inter not bundled w/ matplotlib
    "font.size":         11,
    "legend.frameon":    False,
})


# ===========================================================================
# Constants (defaults)
# ===========================================================================
INDICATORS = [
    "global_debt_to_gdp",
    "growth_slowdown_gap",
    "core_inflation",
    "energy_price_vol",
    "youth_unemployment",
    "tech_displacement_idx",
    "trust_deficit_idx",
    "geopolitical_risk_idx",
    "inequality_gini_gap",
]

INDICATOR_LABELS = {
    "global_debt_to_gdp":    "Global debt / GDP (%)",
    "growth_slowdown_gap":   "Growth slowdown gap (% pts)",
    "core_inflation":        "Core inflation (%)",
    "energy_price_vol":      "Energy price volatility (%, ann.)",
    "youth_unemployment":    "Youth unemployment (%)",
    "tech_displacement_idx": "Tech displacement index (0-100)",
    "trust_deficit_idx":     "Trust deficit index (0-100)",
    "geopolitical_risk_idx": "Geopolitical risk index",
    "inequality_gini_gap":   "Inequality Gini gap (bps)",
}

DEFAULT_READINGS = {
    "global_debt_to_gdp":    305.0,
    "growth_slowdown_gap":     0.7,
    "core_inflation":          3.0,
    "energy_price_vol":       35.0,
    "youth_unemployment":     14.5,
    "tech_displacement_idx":  72.0,
    "trust_deficit_idx":      61.0,
    "geopolitical_risk_idx": 220.0,
    "inequality_gini_gap":     4.5,
}

INDICATOR_RANGES = {
    "global_debt_to_gdp":    (150.0, 450.0, 1.0),
    "growth_slowdown_gap":   (-2.0, 5.0, 0.1),
    "core_inflation":        (0.0, 12.0, 0.1),
    "energy_price_vol":      (5.0, 80.0, 0.5),
    "youth_unemployment":    (4.0, 35.0, 0.1),
    "tech_displacement_idx": (0.0, 100.0, 1.0),
    "trust_deficit_idx":     (0.0, 100.0, 1.0),
    "geopolitical_risk_idx": (40.0, 400.0, 1.0),
    "inequality_gini_gap":   (-5.0, 15.0, 0.1),
}

CALM_BASELINES = {
    "global_debt_to_gdp":   (230.0, 25.0),
    "growth_slowdown_gap":  (0.0, 0.6),
    "core_inflation":       (2.0, 0.7),
    "energy_price_vol":     (22.0, 8.0),
    "youth_unemployment":   (11.5, 2.5),
    "tech_displacement_idx":(50.0, 12.0),
    "trust_deficit_idx":    (45.0, 9.0),
    "geopolitical_risk_idx":(120.0, 45.0),
    "inequality_gini_gap":  (0.0, 2.0),
}


@dataclass
class CountryDebt:
    name: str
    d0: float
    r_base: float
    g_base: float
    pb: float


DEFAULT_COUNTRIES = [
    CountryDebt("United States", 122.0, 4.3, 4.5, -3.0),
    CountryDebt("Japan",         260.0, 1.2, 1.8, -2.0),
    CountryDebt("Italy",         141.0, 3.7, 3.0, -1.0),
    CountryDebt("China",          92.0, 3.0, 5.5, -3.5),
    CountryDebt("India",          82.0, 7.0, 9.5, -2.5),
]

FACTORS = ["growth_shock", "inflation_shock", "oil_shock", "rate_shock", "fragmentation_shock"]
MARGINALS = {
    "growth_shock":        (0.0, 1.0),
    "inflation_shock":     (0.0, 1.0),
    "oil_shock":           (0.0, 25.0),
    "rate_shock":          (0.0, 0.8),
    "fragmentation_shock": (0.0, 0.6),
}
CORR_DEFAULT = np.array([
    [ 1.00, -0.20, -0.10, -0.30, -0.35],
    [-0.20,  1.00,  0.55,  0.40,  0.10],
    [-0.10,  0.55,  1.00,  0.10,  0.20],
    [-0.30,  0.40,  0.10,  1.00,  0.10],
    [-0.35,  0.10,  0.20,  0.10,  1.00],
])

DEFAULT_SCENARIOS = {
    "Base":                     dict(growth_shock=0.0,  inflation_shock=0.0,
                                     oil_shock=0.0,    rate_shock=0.0,
                                     fragmentation_shock=0.0),
    "Slow Burn":                dict(growth_shock=-0.5, inflation_shock=+0.3,
                                     oil_shock=+10.0,  rate_shock=+0.25,
                                     fragmentation_shock=+0.3),
    "Energy Shock + Trade War": dict(growth_shock=-1.5, inflation_shock=+2.5,
                                     oil_shock=+60.0,  rate_shock=+1.0,
                                     fragmentation_shock=+1.2),
}


# ===========================================================================
# Core math — cached
# ===========================================================================
@st.cache_data(show_spinner=False)
def synth_history(readings: dict, seed: int, n_quarters: int = 120) -> pd.DataFrame:
    """Synthesise 30y quarterly indicator history. Cached on readings + seed."""
    rng = np.random.default_rng(seed)
    dates = pd.date_range("1996-Q1", periods=n_quarters, freq="QE")
    df = pd.DataFrame(index=dates, columns=INDICATORS, dtype=float)
    for ind, (mu, sd) in CALM_BASELINES.items():
        x = np.zeros(n_quarters)
        x[0] = mu
        rho = 0.85
        for t in range(1, n_quarters):
            x[t] = mu + rho * (x[t - 1] - mu) + rng.normal(0, sd * np.sqrt(1 - rho ** 2))
        df[ind] = x
    # Crisis shocks
    for q in ["1997-09-30", "2008-09-30", "2020-03-31"]:
        idx_list = df.index.astype(str).tolist()
        if q in idx_list:
            idx = idx_list.index(q)
            for ind, (mu, sd) in CALM_BASELINES.items():
                df.iloc[idx, df.columns.get_loc(ind)] += 2.2 * sd
                for k in range(1, 6):
                    if idx + k < n_quarters:
                        df.iloc[idx + k, df.columns.get_loc(ind)] += (0.6 ** k) * 1.8 * sd
    # Stamp current readings
    for ind, val in readings.items():
        df.iloc[-1, df.columns.get_loc(ind)] = val
    return df


def zscore_panel(df: pd.DataFrame) -> pd.DataFrame:
    return (df - df.mean()) / df.std(ddof=0)


def pca_weights(z: pd.DataFrame) -> np.ndarray:
    C = np.cov(z.values, rowvar=False)
    _, vecs = eigh(C)
    pc1 = vecs[:, -1]
    if pc1.sum() < 0:
        pc1 = -pc1
    return pc1 / np.abs(pc1).sum()


def classify_regime(sri: float) -> str:
    if sri < 25: return "Calm"
    if sri < 50: return "Watch"
    if sri < 75: return "Stress"
    return "Crisis"


@st.cache_data(show_spinner=False)
def compute_sri(readings_tuple: tuple, seed: int):
    readings = dict(zip(INDICATORS, readings_tuple))
    hist = synth_history(readings, seed)
    z = zscore_panel(hist)
    eq_w = np.ones(len(INDICATORS)) / len(INDICATORS)
    pc_w = pca_weights(z)
    raw_eq = z.values @ eq_w
    raw_pc = z.values @ pc_w
    def scale(x):
        s = (x - x.mean()) / x.std(ddof=0)
        return 100.0 * (1.0 / (1.0 + np.exp(-s)))
    sri = pd.DataFrame({
        "sri_equal": scale(raw_eq),
        "sri_pca":   scale(raw_pc),
    }, index=hist.index)
    sri["regime"] = sri["sri_equal"].apply(classify_regime)
    return sri, pd.Series(pc_w, index=INDICATORS, name="pca_weight"), z.iloc[-1]


def debt_dynamics(c: CountryDebt, r: float, g: float, pb: float, horizon=10):
    d = np.zeros(horizon + 1)
    d[0] = c.d0
    factor = (1 + r / 100.0) / (1 + g / 100.0)
    for t in range(1, horizon + 1):
        d[t] = factor * d[t - 1] - pb
    return d


def stabilising_pb(c: CountryDebt, r: float, g: float):
    return ((r - g) / (1 + g / 100.0)) * c.d0 / 100.0


def debt_scenarios(countries, adv, sev, horizon=10):
    scen = {
        "Base":             {"dr":  0.0, "dg":  0.0, "dpb":  0.0},
        "Adverse":          adv,
        "Severely Adverse": sev,
    }
    out = {}
    for c in countries:
        out[c.name] = {}
        for sname, s in scen.items():
            r = c.r_base + s["dr"]
            g = c.g_base + s["dg"]
            pb = c.pb + s["dpb"]
            path = debt_dynamics(c, r, g, pb, horizon=horizon)
            out[c.name][sname] = {
                "path": path,
                "pb_star": stabilising_pb(c, r, g),
                "pb_gap": stabilising_pb(c, r, g) - pb,
                "d_terminal": path[-1],
                "r": r, "g": g, "pb": pb,
            }
    return out


@st.cache_data(show_spinner=False)
def mc_stress_test(n: int, seed: int, baseline_growth: float,
                   corr_flat: tuple) -> dict:
    corr = np.array(corr_flat).reshape(len(FACTORS), len(FACTORS))
    rng = np.random.default_rng(seed)
    L = np.linalg.cholesky(corr)
    z = rng.standard_normal((n, len(FACTORS))) @ L.T
    factors = {}
    for i, f in enumerate(FACTORS):
        mu, sd = MARGINALS[f]
        factors[f] = mu + sd * z[:, i]
    f = pd.DataFrame(factors)
    g = (
        baseline_growth
        + 1.0 * f["growth_shock"]
        - 0.20 * f["inflation_shock"]
        - 0.015 * f["oil_shock"]
        - 0.40 * f["rate_shock"]
        - 0.60 * f["fragmentation_shock"]
    ).rename("growth_1y")
    var95 = float(np.percentile(g, 5))
    cvar95 = float(g[g <= var95].mean())
    return {
        "n_paths":      n,
        "mean_growth":  float(g.mean()),
        "std_growth":   float(g.std(ddof=0)),
        "var95":        var95,
        "cvar95":       cvar95,
        "p_recession":  float((g < 0).mean()),
        "growth":       g,
        "factors":      f,
    }


def scenario_outcome(shocks: dict, baseline_growth: float = 3.1) -> dict:
    g = (
        baseline_growth
        + 1.0 * shocks["growth_shock"]
        - 0.20 * shocks["inflation_shock"]
        - 0.015 * shocks["oil_shock"]
        - 0.40 * shocks["rate_shock"]
        - 0.60 * shocks["fragmentation_shock"]
    )
    infl = 3.0 + 1.0 * shocks["inflation_shock"] + 0.02 * shocks["oil_shock"]
    equity_dd = (
        -2.0
        + 6.0 * (-shocks["growth_shock"])
        + 4.0 * shocks["fragmentation_shock"]
        + 0.2 * shocks["oil_shock"]
    )
    return {
        "growth_1y": float(g),
        "inflation_1y": float(infl),
        "equity_dd_1y": float(-equity_dd),
    }


# ===========================================================================
# Charts
# ===========================================================================
def fig_sri(sri: pd.DataFrame):
    fig, ax = plt.subplots(figsize=(10, 4.5))
    ax.plot(sri.index, sri["sri_equal"], label="SRI (equal weight)",
            lw=2, color=NAVY)
    ax.plot(sri.index, sri["sri_pca"],   label="SRI (PCA)",
            lw=2, alpha=0.9, color=GOLD)
    for y, label, color in [(25, "Calm",   MOUNTAIN_GREEN),
                            (50, "Watch",  ALERT_AMBER),
                            (75, "Stress", ALERT_RED)]:
        ax.axhline(y, color=color, ls="--", lw=1, alpha=0.7)
        ax.text(sri.index[3], y + 1, label, color=color, fontsize=9)
    ax.set_title("Composite Systemic Risk Index — 1996 to 2026")
    ax.set_ylabel("SRI (0 = calm, 100 = crisis)")
    ax.set_ylim(0, 100)
    ax.legend(loc="upper left")
    fig.tight_layout()
    return fig


def fig_zscores(z_now: pd.Series):
    fig, ax = plt.subplots(figsize=(10, 4.5))
    z = z_now.reindex(INDICATORS)
    colors = [ALERT_RED if v > 1.5 else (GOLD if v > 0 else NAVY) for v in z]
    ax.barh(range(len(z)), z.values, color=colors, edgecolor=NAVY_DEEP, linewidth=0.4)
    ax.set_yticks(range(len(z)))
    ax.set_yticklabels([INDICATOR_LABELS[i] for i in z.index])
    ax.axvline(0, color=NAVY_DEEP, lw=0.8)
    ax.axvline(1, color=ALERT_AMBER, ls="--", lw=1, alpha=0.7)
    ax.axvline(2, color=ALERT_RED,   ls="--", lw=1, alpha=0.7)
    ax.set_xlabel("Z-score vs 30y history")
    ax.set_title("Current indicator z-scores")
    ax.invert_yaxis()
    fig.tight_layout()
    return fig


def fig_debt_fan(results, countries):
    fig, axes = plt.subplots(1, len(countries), figsize=(16, 4.5), sharey=False)
    if len(countries) == 1:
        axes = [axes]
    colors = {"Base": NAVY, "Adverse": GOLD, "Severely Adverse": ALERT_RED}
    for ax, c in zip(axes, countries):
        for sname, color in colors.items():
            path = results[c.name][sname]["path"]
            ax.plot(range(len(path)), path, label=sname, color=color, lw=2)
        ax.set_title(c.name, fontsize=11)
        ax.set_xlabel("Years")
        ax.set_ylabel("Debt / GDP (%)")
        ax.legend(fontsize=8)
    fig.suptitle("Debt sustainability — 10y horizon", fontsize=13)
    fig.tight_layout()
    return fig


def fig_mc(mc):
    g = mc["growth"]
    fig, ax = plt.subplots(figsize=(10, 4.5))
    ax.hist(g, bins=60, color=NAVY, alpha=0.88, edgecolor=PARCHMENT)
    ax.axvline(mc["var95"], color=ALERT_RED, ls="--", lw=2,
               label=f"VaR 95%: {mc['var95']:.2f}%")
    ax.axvline(mc["cvar95"], color=GOLD_DARK, ls="--", lw=2,
               label=f"CVaR 95%: {mc['cvar95']:.2f}%")
    ax.axvline(0, color=NAVY_DEEP, ls=":", lw=1.5, label="Recession line")
    ax.set_title(f"Monte Carlo 1y global growth (n={mc['n_paths']:,})")
    ax.set_xlabel("Global growth, next 12 months (%)")
    ax.set_ylabel("Count")
    ax.legend()
    fig.tight_layout()
    return fig


def fig_scenarios(scen_results: dict):
    labels = list(scen_results.keys())
    growths = [scen_results[s]["growth_1y"] for s in labels]
    infls   = [scen_results[s]["inflation_1y"] for s in labels]
    dds     = [scen_results[s]["equity_dd_1y"] for s in labels]
    fig, axes = plt.subplots(1, 3, figsize=(15, 4.2))
    colors = [NAVY, GOLD, ALERT_RED]
    for ax, vals, title, ylabel in zip(
        axes,
        [growths, infls, dds],
        ["1y growth", "1y inflation", "Equity drawdown"],
        ["%", "%", "%"],
    ):
        ax.bar(labels, vals, color=colors[: len(labels)],
               edgecolor=NAVY_DEEP, linewidth=0.5)
        ax.set_title(title)
        ax.set_ylabel(ylabel)
        ax.tick_params(axis="x", rotation=15)
        for i, v in enumerate(vals):
            ax.text(i, v, f"{v:+.1f}", ha="center",
                    va="bottom" if v >= 0 else "top",
                    fontsize=10, color=NAVY_DEEP, fontweight="bold")
    fig.tight_layout()
    return fig


# ===========================================================================
# Excel dashboard builder — returns bytes
# ===========================================================================
def build_excel_bytes(summary: dict) -> bytes:
    ACCENT_HEX = "1F4E79"
    thin = Side(style="thin", color="BFBFBF")
    box = Border(top=thin, bottom=thin, left=thin, right=thin)
    WHITE_BOLD  = Font(bold=True, color="FFFFFF", name="Calibri", size=11)
    ACCENT_BOLD = Font(bold=True, color=ACCENT_HEX, name="Calibri", size=12)
    TITLE_FONT  = Font(bold=True, color=ACCENT_HEX, name="Calibri", size=18)
    SUB_FONT    = Font(italic=True, color="595959", name="Calibri", size=11)
    BODY        = Font(name="Calibri", size=11)
    BODY_BOLD   = Font(name="Calibri", size=11, bold=True)

    def hdr_fill():    return PatternFill("solid", fgColor=ACCENT_HEX)
    def stripe_fill(): return PatternFill("solid", fgColor="F4F6FB")
    def soft_fill(c):  return PatternFill("solid", fgColor=c)

    def hrow(ws, row, headers, start_col=1):
        for i, h in enumerate(headers):
            c = ws.cell(row=row, column=start_col + i, value=h)
            c.font, c.fill, c.border = WHITE_BOLD, hdr_fill(), box
            c.alignment = Alignment(horizontal="left", vertical="center")

    def brow(ws, row, values, bold_first=True, start_col=1, stripe=False):
        for i, v in enumerate(values):
            c = ws.cell(row=row, column=start_col + i, value=v)
            c.font = BODY_BOLD if (bold_first and i == 0) else BODY
            c.alignment = Alignment(horizontal="left", vertical="center")
            c.border = box
            if stripe:
                c.fill = stripe_fill()

    def widths(ws, ws_widths):
        for i, w in enumerate(ws_widths, 1):
            ws.column_dimensions[get_column_letter(i)].width = w

    def title(ws, t, sub):
        ws.cell(row=1, column=1, value=t).font = TITLE_FONT
        ws.cell(row=2, column=1, value=sub).font = SUB_FONT
        ws.row_dimensions[1].height = 26

    wb = Workbook(); wb.remove(wb.active)

    # Overview
    ws = wb.create_sheet("Overview", 0)
    widths(ws, [4, 30, 22, 22, 22, 22])
    title(ws, "Macro Risk Dashboard",
          f"As of {summary['as_of']} · interactive build")
    ws.cell(row=4, column=2, value="Headline").font = ACCENT_BOLD
    rows = [
        ("Composite SRI (equal weight)", summary["sri"]["current_equal_weight"], summary["sri"]["regime"]),
        ("Composite SRI (PCA weight)",   summary["sri"]["current_pca"], ""),
        ("MC mean 1y global growth",     f"{summary['monte_carlo']['mean_growth']}%", ""),
        ("MC VaR 95%",                   f"{summary['monte_carlo']['var95']}%", ""),
        ("MC CVaR 95%",                  f"{summary['monte_carlo']['cvar95']}%", ""),
        ("P(global recession, 1y)",      f"{summary['monte_carlo']['p_recession'] * 100:.1f}%", ""),
    ]
    hrow(ws, 5, ["Metric", "Value", "Regime"], start_col=2)
    for i, (lbl, val, reg) in enumerate(rows):
        brow(ws, 6 + i, [lbl, val, reg], start_col=2, stripe=(i % 2 == 0))
    color_map = {"Calm": "C6E0B4", "Watch": "FFE699",
                 "Stress": "F8CBAD", "Crisis": "F4B084"}
    ws.cell(row=6, column=4).fill = soft_fill(color_map.get(summary["sri"]["regime"], "FFFFFF"))
    ws.cell(row=6, column=4).font = BODY_BOLD

    ws.cell(row=14, column=2, value="Scenario engine — 1-year outcomes").font = ACCENT_BOLD
    hrow(ws, 15, ["Scenario", "Growth (%)", "Inflation (%)", "Equity drawdown (%)"], start_col=2)
    for i, (n, v) in enumerate(summary["scenarios"].items()):
        brow(ws, 16 + i,
             [n, v["growth_1y"], v["inflation_1y"], v["equity_dd_1y"]],
             start_col=2, stripe=(i % 2 == 0))

    # Debt
    ws = wb.create_sheet("Debt")
    widths(ws, [4, 18, 16, 22, 22, 22])
    title(ws, "Debt Sustainability Dynamics", "Terminal debt/GDP and PB gap by scenario")
    hrow(ws, 4, ["Country", "Scenario", "Terminal debt/GDP (%)",
                 "Stabilising PB (% GDP)", "Required PB gap (% GDP)"], start_col=2)
    r = 5
    for country, scens in summary["debt"].items():
        for sn, vals in scens.items():
            brow(ws, r, [country, sn,
                         vals["d_terminal_pct"], vals["pb_star_pct"], vals["pb_gap_pct"]],
                 start_col=2, stripe=((r - 5) % 2 == 0))
            r += 1
    ws.conditional_formatting.add(
        f"F5:F{r - 1}",
        ColorScaleRule(start_type="num", start_value=0, start_color="C6E0B4",
                       mid_type="num", mid_value=5, mid_color="FFE699",
                       end_type="num", end_value=12, end_color="F4B084"),
    )

    # MC sheet
    ws = wb.create_sheet("MonteCarlo")
    widths(ws, [4, 30, 16])
    title(ws, "Monte Carlo Stress Test", "Gaussian copula on 5 correlated factors")
    hrow(ws, 4, ["Metric", "Value"], start_col=2)
    mc = summary["monte_carlo"]
    metrics = [
        ("Paths", mc["n_paths"]),
        ("Mean 1y growth (%)", mc["mean_growth"]),
        ("Std dev 1y growth (%)", mc["std_growth"]),
        ("VaR 95% (%)", mc["var95"]),
        ("CVaR 95% (%)", mc["cvar95"]),
        ("P(global recession, 1y)", mc["p_recession"]),
    ]
    for i, (k, v) in enumerate(metrics):
        brow(ws, 5 + i, [k, v], start_col=2, stripe=(i % 2 == 0))

    # Scenarios sheet
    ws = wb.create_sheet("Scenarios")
    widths(ws, [4, 28, 16, 18, 22])
    title(ws, "Scenario Engine", "Single-shot reduced-form macro outcomes")
    hrow(ws, 4, ["Scenario", "Growth (%)", "Inflation (%)", "Equity drawdown (%)"], start_col=2)
    for i, (n, v) in enumerate(summary["scenarios"].items()):
        brow(ws, 5 + i,
             [n, v["growth_1y"], v["inflation_1y"], v["equity_dd_1y"]],
             start_col=2, stripe=(i % 2 == 0))

    buf = io.BytesIO()
    wb.save(buf)
    buf.seek(0)
    return buf.read()


# ===========================================================================
# Sidebar — branded header + controls
# ===========================================================================
LINKEDIN_URL = "https://www.linkedin.com/in/trichyravis"
GITHUB_URL   = "https://github.com/trichyravis/"
ACADEMY_URL  = "https://themountainpathacademy.com"

st.sidebar.markdown(
    f"""
    <div class="mpa-profile">
        <div style="font-size:1.8rem;line-height:1">⛰</div>
        <div class="name">The Mountain Path</div>
        <div class="tagline">Finance · Risk · Modelling</div>
        <a href="{LINKEDIN_URL}" target="_blank">in · LinkedIn</a>
        <a href="{GITHUB_URL}" target="_blank">⌥ GitHub</a>
        <div style="margin-top:0.5rem">
            <a href="{ACADEMY_URL}" target="_blank">⛰ Academy</a>
        </div>
    </div>
    """,
    unsafe_allow_html=True,
)
st.sidebar.markdown("### Macro Risk Controls")
st.sidebar.caption("Adjust any input; the dashboard recomputes live.")

with st.sidebar.expander("Global settings", expanded=True):
    seed = st.number_input("Random seed", value=20260519, step=1)
    n_paths = st.select_slider("Monte Carlo paths",
                               options=[1000, 2500, 5000, 10000, 25000, 50000],
                               value=10000)
    baseline_growth = st.slider("Baseline 1y global growth (%)",
                                0.0, 6.0, 3.1, 0.1)

with st.sidebar.expander("Current indicator readings", expanded=False):
    readings = {}
    for ind in INDICATORS:
        lo, hi, step = INDICATOR_RANGES[ind]
        readings[ind] = st.slider(
            INDICATOR_LABELS[ind],
            min_value=float(lo), max_value=float(hi),
            value=float(DEFAULT_READINGS[ind]), step=float(step),
        )

with st.sidebar.expander("Debt — country parameters", expanded=False):
    countries = []
    for c in DEFAULT_COUNTRIES:
        st.markdown(f"**{c.name}**")
        d0  = st.slider(f"{c.name} · debt/GDP",  20.0, 400.0, float(c.d0),  1.0, key=f"d0_{c.name}")
        rb  = st.slider(f"{c.name} · r (%)",      0.0, 12.0, float(c.r_base), 0.1, key=f"r_{c.name}")
        gb  = st.slider(f"{c.name} · g (%)",     -2.0, 14.0, float(c.g_base), 0.1, key=f"g_{c.name}")
        pb  = st.slider(f"{c.name} · primary balance (% GDP)",
                        -8.0, 8.0, float(c.pb), 0.1, key=f"pb_{c.name}")
        countries.append(CountryDebt(c.name, d0, rb, gb, pb))

with st.sidebar.expander("Debt — scenario shocks", expanded=False):
    st.caption("Shifts applied on top of each country's base r / g / pb.")
    st.markdown("**Adverse**")
    adv = {
        "dr":  st.slider("Adv · Δr (% pts)",   0.0, 5.0, 1.5, 0.1),
        "dg":  st.slider("Adv · Δg (% pts)",  -5.0, 0.0, -1.0, 0.1),
        "dpb": st.slider("Adv · Δpb (% pts)", -5.0, 2.0, -1.0, 0.1),
    }
    st.markdown("**Severely Adverse**")
    sev = {
        "dr":  st.slider("Sev · Δr (% pts)",   0.0, 8.0, 3.0, 0.1),
        "dg":  st.slider("Sev · Δg (% pts)",  -8.0, 0.0, -2.5, 0.1),
        "dpb": st.slider("Sev · Δpb (% pts)", -8.0, 2.0, -2.5, 0.1),
    }

with st.sidebar.expander("Scenario engine — shocks", expanded=False):
    scenarios = {}
    for name in DEFAULT_SCENARIOS:
        st.markdown(f"**{name}**")
        d = DEFAULT_SCENARIOS[name]
        scenarios[name] = dict(
            growth_shock=        st.slider(f"{name} · growth shock",   -3.0, 3.0,
                                           float(d["growth_shock"]), 0.1, key=f"gs_{name}"),
            inflation_shock=     st.slider(f"{name} · inflation shock", -2.0, 5.0,
                                           float(d["inflation_shock"]), 0.1, key=f"is_{name}"),
            oil_shock=           st.slider(f"{name} · oil shock (%)",  -50.0, 120.0,
                                           float(d["oil_shock"]), 1.0, key=f"os_{name}"),
            rate_shock=          st.slider(f"{name} · rate shock",     -1.0, 3.0,
                                           float(d["rate_shock"]), 0.05, key=f"rs_{name}"),
            fragmentation_shock= st.slider(f"{name} · fragmentation",  -1.0, 2.0,
                                           float(d["fragmentation_shock"]), 0.05, key=f"fs_{name}"),
        )


# ===========================================================================
# Compute
# ===========================================================================
readings_tuple = tuple(readings[i] for i in INDICATORS)
sri, pc_w, z_now = compute_sri(readings_tuple, seed)
debt_res = debt_scenarios(countries, adv, sev)
mc = mc_stress_test(int(n_paths), int(seed), baseline_growth,
                    tuple(CORR_DEFAULT.flatten().tolist()))
scen_results = {n: scenario_outcome(s, baseline_growth) for n, s in scenarios.items()}

current_sri_eq = float(sri["sri_equal"].iloc[-1])
current_sri_pc = float(sri["sri_pca"].iloc[-1])
regime = sri["regime"].iloc[-1]


# ===========================================================================
# Branded header
# ===========================================================================
st.markdown(
    f"""
    <div class="mpa-header">
        <div class="eyebrow">The Mountain Path Academy · Macro Risk Lab</div>
        <h1>Is an Extreme World Economic Collapse Impending?</h1>
        <div class="sub">A practitioner-led macro risk dashboard — debt, growth, energy, geopolitics.</div>
        <div class="by">Interactive companion to the May 2026 briefing · Adjust any input in the sidebar to recompute live.</div>
    </div>
    """,
    unsafe_allow_html=True,
)

col1, col2, col3, col4, col5 = st.columns(5)
regime_color = {"Calm": "#7CB342", "Watch": "#FFB300",
                "Stress": "#E64A19", "Crisis": "#C0392B"}.get(regime, "#000")
col1.metric("SRI (equal weight)", f"{current_sri_eq:.1f}")
col2.metric("SRI (PCA)", f"{current_sri_pc:.1f}")
col3.markdown(f"**Regime**<br><span style='color:{regime_color};"
              f"font-size:1.4em;font-weight:bold'>{regime}</span>",
              unsafe_allow_html=True)
col4.metric("MC VaR 95%", f"{mc['var95']:.2f}%")
col5.metric("P(recession, 1y)", f"{mc['p_recession'] * 100:.1f}%")

st.markdown("---")


# ===========================================================================
# Tabs
# ===========================================================================
tab_over, tab_sri, tab_debt, tab_mc, tab_scen, tab_dl = st.tabs(
    ["Overview", "SRI", "Debt", "Monte Carlo", "Scenarios", "Download"]
)

# ---------- Overview ----------
with tab_over:
    left, right = st.columns([3, 2])
    with left:
        st.subheader("Composite Systemic Risk Index")
        st.pyplot(fig_sri(sri), width="stretch")
    with right:
        st.subheader("Scenarios — 1y outcomes")
        scen_df = pd.DataFrame(scen_results).T
        scen_df.columns = ["Growth (%)", "Inflation (%)", "Equity DD (%)"]
        st.dataframe(scen_df.style.format("{:+.2f}").background_gradient(
            cmap="RdYlGn", axis=0, subset=["Growth (%)"]
        ), width="stretch")
        st.subheader("Monte Carlo — tail metrics")
        mc_df = pd.DataFrame({
            "Metric": ["Mean 1y growth", "VaR 95%", "CVaR 95%", "P(recession)"],
            "Value":  [f"{mc['mean_growth']:.2f}%",
                       f"{mc['var95']:.2f}%",
                       f"{mc['cvar95']:.2f}%",
                       f"{mc['p_recession'] * 100:.1f}%"],
        })
        st.dataframe(mc_df, width="stretch", hide_index=True)


# ---------- SRI ----------
with tab_sri:
    st.subheader("Composite Systemic Risk Index")
    st.pyplot(fig_sri(sri), width="stretch")

    st.subheader("Current indicator z-scores")
    st.caption("Higher = riskier vs 30y history. Z > 2 is a strong tail signal.")
    st.pyplot(fig_zscores(z_now), width="stretch")

    st.subheader("PCA weights vs current readings")
    table = pd.DataFrame({
        "Indicator":     [INDICATOR_LABELS[i] for i in INDICATORS],
        "Current value": [readings[i] for i in INDICATORS],
        "Z-score":       [z_now[i] for i in INDICATORS],
        "PCA weight":    [pc_w[i] for i in INDICATORS],
    })
    st.dataframe(
        table.style.format({"Current value": "{:.2f}",
                            "Z-score": "{:+.2f}",
                            "PCA weight": "{:.3f}"})
        .background_gradient(cmap="OrRd", subset=["Z-score"])
        .background_gradient(cmap="Blues", subset=["PCA weight"]),
        width="stretch", hide_index=True,
    )


# ---------- Debt ----------
with tab_debt:
    st.subheader("Debt sustainability — 10y horizon")
    st.pyplot(fig_debt_fan(debt_res, countries), width="stretch")

    rows = []
    for cname, scens in debt_res.items():
        for sn, v in scens.items():
            rows.append({
                "Country": cname,
                "Scenario": sn,
                "r (%)":  v["r"],
                "g (%)":  v["g"],
                "pb (%)": v["pb"],
                "Terminal debt/GDP (%)": round(v["d_terminal"], 1),
                "Stabilising PB (% GDP)": round(v["pb_star"], 2),
                "Required PB gap (% GDP)": round(v["pb_gap"], 2),
            })
    debt_df = pd.DataFrame(rows)
    st.dataframe(
        debt_df.style.format({
            "r (%)": "{:.2f}", "g (%)": "{:.2f}", "pb (%)": "{:+.2f}",
            "Terminal debt/GDP (%)": "{:.1f}",
            "Stabilising PB (% GDP)": "{:+.2f}",
            "Required PB gap (% GDP)": "{:+.2f}",
        }).background_gradient(cmap="Reds", subset=["Required PB gap (% GDP)"]),
        width="stretch", hide_index=True,
    )

    st.caption(
        "The required primary-balance gap = the tightening (% of GDP) needed "
        "to stabilise debt/GDP at the current level under each scenario's r and g."
    )


# ---------- Monte Carlo ----------
with tab_mc:
    st.subheader("Monte Carlo 1y global growth distribution")
    st.pyplot(fig_mc(mc), width="stretch")

    a, b, c, d = st.columns(4)
    a.metric("Mean growth", f"{mc['mean_growth']:.2f}%")
    b.metric("Std dev",     f"{mc['std_growth']:.2f}%")
    c.metric("VaR 95%",     f"{mc['var95']:.2f}%")
    d.metric("CVaR 95%",    f"{mc['cvar95']:.2f}%")
    st.metric("P(global recession, 1y)", f"{mc['p_recession'] * 100:.1f}%")

    with st.expander("Factor correlation matrix"):
        corr_df = pd.DataFrame(CORR_DEFAULT, index=FACTORS, columns=FACTORS)
        st.dataframe(
            corr_df.style.format("{:.2f}").background_gradient(
                cmap="RdBu_r", vmin=-1, vmax=1
            ),
            width="stretch",
        )

    with st.expander("Factor sample (first 1,000 paths)"):
        st.dataframe(mc["factors"].head(1000), width="stretch")


# ---------- Scenarios ----------
with tab_scen:
    st.subheader("Scenario engine — single-shot outcomes")
    st.pyplot(fig_scenarios(scen_results), width="stretch")

    st.subheader("Outcomes table")
    df = pd.DataFrame(scen_results).T
    df.columns = ["Growth (%)", "Inflation (%)", "Equity drawdown (%)"]
    st.dataframe(
        df.style.format("{:+.2f}")
          .background_gradient(cmap="RdYlGn", subset=["Growth (%)"])
          .background_gradient(cmap="OrRd",  subset=["Inflation (%)"]),
        width="stretch",
    )

    st.subheader("Shock inputs")
    st.dataframe(
        pd.DataFrame(scenarios).T,
        width="stretch",
    )


# ---------- Download ----------
with tab_dl:
    st.subheader("Export")
    st.write("Generate an Excel dashboard or a JSON dump of the current run.")

    summary = {
        "as_of": "2026-05-19",
        "sri": {
            "current_equal_weight": round(current_sri_eq, 1),
            "current_pca": round(current_sri_pc, 1),
            "regime": regime,
            "pca_weights": {k: round(float(v), 3) for k, v in pc_w.items()},
            "current_readings": readings,
        },
        "debt": {
            cname: {
                sn: {
                    "d_terminal_pct": round(v["d_terminal"], 1),
                    "pb_star_pct":    round(v["pb_star"], 2),
                    "pb_gap_pct":     round(v["pb_gap"], 2),
                } for sn, v in scens.items()
            } for cname, scens in debt_res.items()
        },
        "monte_carlo": {
            "n_paths":     mc["n_paths"],
            "mean_growth": round(mc["mean_growth"], 2),
            "std_growth":  round(mc["std_growth"], 2),
            "var95":       round(mc["var95"], 2),
            "cvar95":      round(mc["cvar95"], 2),
            "p_recession": round(mc["p_recession"], 3),
        },
        "scenarios": {
            n: {k: round(v, 2) for k, v in d.items()}
            for n, d in scen_results.items()
        },
    }

    xlsx_bytes = build_excel_bytes(summary)
    col_x, col_j = st.columns(2)
    col_x.download_button(
        "⬇ Download risk_dashboard.xlsx",
        data=xlsx_bytes,
        file_name="risk_dashboard.xlsx",
        mime="application/vnd.openxmlformats-officedocument.spreadsheetml.sheet",
        width="stretch",
    )
    col_j.download_button(
        "⬇ Download summary.json",
        data=json.dumps(summary, indent=2).encode(),
        file_name="summary.json",
        mime="application/json",
        width="stretch",
    )

    with st.expander("Preview JSON"):
        st.json(summary)


st.markdown(
    f"""
    <div class="mpa-footer">
        <div class="col">
            <h4>The Mountain Path Academy</h4>
            <a href="{ACADEMY_URL}" target="_blank">themountainpathacademy.com</a><br>
            <span style="color:{SLATE};font-size:0.85rem">Practitioner-led courses in financial<br>
            modelling, risk, derivatives & valuation</span>
        </div>
        <div class="col">
            <h4>Connect</h4>
            <a href="{LINKEDIN_URL}" target="_blank">LinkedIn → trichyravis</a><br>
            <a href="{GITHUB_URL}" target="_blank">GitHub → trichyravis</a>
        </div>
        <div class="col">
            <h4>Methodology</h4>
            <span style="color:{SLATE};font-size:0.85rem">
            9-indicator composite SRI · debt sustainability<br>
            simulation · 10k-path Gaussian copula MC · scenarios
            </span>
        </div>
        <div class="disclaimer">
            Illustrative model — calibrated to the May 2026 briefing. Replace
            <code>synth_history()</code> with real IMF / World Bank / FRED pulls to productionise.
            Not investment advice. © {2026} V. Ravichandran · The Mountain Path Academy.
        </div>
    </div>
    """,
    unsafe_allow_html=True,
)
